#!/usr/bin/env python

########
# summarize_tests.py
#
# Robby Stahl, F5 Networks (r.stahl@f5.com)
# 2015-2016
#
# This program traverses the directory tree from CWD in search of .xlsx
# files. It assumes that any .xlsx files that it encounters are output
# from ltmStats.pl. It will non-destructively read the contents of
# these files, and write the output in CWD/summarized.xlsx.
########


# Excel reading
from openpyxl import reader, load_workbook, Workbook

# statistics
import numpy

# used for directory traversal / file listing
import os


#### <constants> ####
#DEBUG = 1
DEBUG = 0

DATA_SHEET_NAME = 'raw_data'
OUTPUT_WORKBOOK = 'summarized.xlsx'
#### </constants> ####


class MyStats:
	def __init__(self):
		self.val_mean = 0.0
		self.val_min = 0.0
		self.val_max = 0.0
		self.val_stdev = 0.0


class PerfData:
	"""
	Instances of this class contain the raw data and summary statistics from a sheet.
	"""
	def __init__(self):
		self.run_time = []
		self.sys_cpu = []
		self.tmm_cpu = []
		self.memory = []
		self.client_bytes_in = []
		self.client_bytes_out = []
		self.server_bytes_in = []
		self.server_bytes_out = []
		self.client_conn_active = []
		self.client_conn_count = []
		self.server_conn_active = []
		self.server_conn_count = []

		# self.run_time_stats = MyStats()
		self.sys_cpu_stats = MyStats()
		self.tmm_cpu_stats = MyStats()
		self.memory_stats = MyStats()
		self.client_mb_in_stats = MyStats()
		self.client_mb_out_stats = MyStats()
		self.server_mb_in_stats = MyStats()
		self.server_mb_out_stats = MyStats()
		self.client_conn_active_stats = MyStats()
		self.client_conn_count_stats = MyStats()
		self.server_conn_active_stats = MyStats()
		self.server_conn_count_stats = MyStats()


def summarize_raw(vals):
	"""
	Takes a list of values, return a populated MyStats object.
	"""
	# Accumulator-based columns use deltas, thereby prevening the first item
	# from appearing in the output. Do the same for consistency's sake.
	
	a = vals[:]
	a.pop(0) # bye bye, bits!
	stats = MyStats()
	stats.val_mean = numpy.mean(a)
	stats.val_min = min(a)
	stats.val_max = max(a)
	stats.val_stdev = numpy.std(a)
	
	return stats
	

def summarize_memory(vals):
	"""
	Takes a list of BIG-IP SNMP memory counts, returns a populated MyStats object.
	Values are in MB, base 2 (1024 * 1024)
	"""
	unit = (1024 * 1024 * 1.0) # bytes to megabytes, force floating point
	# Accumulator-based columns use deltas, thereby prevening the first item
	# from appearing in the output. Do the same for consistency's sake.

	a = vals[:]
	a.pop(0) # bye bye, bits!
	if (DEBUG):
		print "summarize_memory(...) size of a[] " + str(len(a))

	for i in xrange(len(a)):
		a[i] = a[i] / unit

	stats = MyStats()
	stats.val_mean = numpy.mean(a)
	stats.val_min = min(a)
	stats.val_max = max(a)
	stats.val_stdev = numpy.std(a)

	return stats


def summarize_throughput(vals, times):
	"""
	Takes two lists (BIG-IP SNMP bytes accumulator, timestamps), returns a populated
	MyStats object. Raw values are in bytes/s, summary results are in Mbps.
	"""
	unit = ((1000 * 1000) / 8.0) # bytes to megabits, force floating point

	a = vals[:]
	t = times[:]
	r = [] # calculated results
	for i in xrange(1, len(a)): # starting at 1, because we're finding deltas
		r.append( ((a[i] - a[(i - 1)]) / (t[i] - t[(i - 1)])) / unit)
	#	if (DEBUG):
	#		print i, a[i], a[i-1], t[i], t[i-1]
	
	if (DEBUG):
		print "summarize_throughput(...) size of r[] " + str(len(r))
	
	stats = MyStats()
	stats.val_mean = numpy.mean(r)
	stats.val_min = min(r)
	stats.val_max = max(r)
	stats.val_stdev = numpy.std(r)

	return stats


def summarize_counts(vals, times):
	"""
	Takes two lists (BIG-IP SNMP count accumulator, timestamps), returns a populated
	MyStats object. Raw values are unbounded integers.
	"""

	a = vals[:]
	t = times[:]
	r = [] # calculated results
	for i in xrange(1, len(a)): # starting at 1, because we're finding deltas
		r.append((a[i] - a[(i - 1)]) / (t[i] - t[(i - 1)])) 
	
	stats = MyStats()
	stats.val_mean = numpy.mean(r)
	stats.val_min = min(r)
	stats.val_max = max(r)
	stats.val_stdev = numpy.std(r)

	return stats


def print_stats(stats, title=None):
	"""
	Prints the contents of a MyStats object. Will use a provided title.
	"""
	if (title):
		print "---- " + title + " ----"
	print "mean:  " + str(stats.val_mean)
	print "min:   " + str(stats.val_min)
	print "max:   " + str(stats.val_max)
	print "stdev: " + str(stats.val_stdev)
	print # trailing newline


def select_workbooks():
	"""
	Takes no arguments. Traverses the directory structure from current, builds a list
	of .xlsx spreadsheets. Returns the constructed list.
	"""
	# walk the tree, retain everything that ends in .xlsx
	# TODO: check that the file actually contains an Excel workbook
	to_return = []
	for myroot, dirs, files in os.walk("."):
		for f in files:
			f_temp = os.path.join(myroot, f) # f_temp is now the relative path + filename
			if (f_temp.endswith('.xlsx')):
				to_return.append(f_temp)
	return to_return


def process_data(workbook_fname):
	"""
	Takes a workbook filename, summarizes the raw data contained within. Returns a PerfData object.
	"""
	to_return = PerfData()
	wb = load_workbook(workbook_fname)
	ws = wb.get_sheet_by_name(DATA_SHEET_NAME) # DATA_SHEET_NAME is defined near the top of this file
	max_row = ws.get_highest_row()

	if (DEBUG):
		print "process_data(...) value of max_row " + str(max_row)

	# grab data from the raw data sheet
	for i in xrange(1,(max_row)): # skip column names
		# excel is 1-indexed, xrange is 0-indexed. This is the reason for the + 1 in each seek.
		to_return.run_time.append(ws['A'+str(i + 1)].value)
		to_return.sys_cpu.append(ws['B'+str(i + 1)].value)
		to_return.tmm_cpu.append(ws['C'+str(i + 1)].value)
		to_return.memory.append(ws['D'+str(i + 1)].value)
		to_return.client_bytes_in.append(ws['E'+str(i + 1)].value)
		to_return.client_bytes_out.append(ws['F'+str(i + 1)].value)
		to_return.server_bytes_in.append(ws['I'+str(i + 1)].value)
		to_return.server_bytes_out.append(ws['J'+str(i + 1)].value)
		to_return.client_conn_active.append(ws['M'+str(i + 1)].value)
		to_return.client_conn_count.append(ws['N'+str(i + 1)].value)
		to_return.server_conn_active.append(ws['O'+str(i + 1)].value)
		to_return.server_conn_count.append(ws['P'+str(i + 1)].value)
	
	if (DEBUG):
		print "process_data(...) size of to_return.run_time " + str(len(to_return.run_time))
	
	# generate summary statistics
	to_return.sys_cpu_stats = summarize_raw(to_return.sys_cpu)
	to_return.tmm_cpu_stats = summarize_raw(to_return.tmm_cpu)
	to_return.memory_stats = summarize_memory(to_return.memory)
	to_return.client_mb_in_stats = summarize_throughput(to_return.client_bytes_in, to_return.run_time)
	to_return.client_mb_out_stats = summarize_throughput(to_return.client_bytes_out, to_return.run_time)
	to_return.server_mb_in_stats = summarize_throughput(to_return.server_bytes_in, to_return.run_time)
	to_return.server_mb_out_stats = summarize_throughput(to_return.server_bytes_out, to_return.run_time)
	to_return.client_conn_active_stats = summarize_raw(to_return.client_conn_active)
	to_return.client_conn_count_stats = summarize_counts(to_return.client_conn_count, to_return.run_time)
	to_return.server_conn_active_stats = summarize_raw(to_return.server_conn_active)
	to_return.server_conn_count_stats = summarize_counts(to_return.server_conn_count, to_return.run_time)
	
	# let's see what it looks like
	if (DEBUG):
		print_stats(to_return.sys_cpu_stats, "System CPU")
		print_stats(to_return.tmm_cpu_stats, "TMM CPU")
		print_stats(to_return.memory_stats, "Memory")
		print_stats(to_return.client_mb_in_stats, "Client Mbps in")
		print_stats(to_return.client_mb_out_stats, "Client Mbps out")
		print_stats(to_return.server_mb_in_stats, "Server Mbps in")
		print_stats(to_return.server_mb_out_stats, "Server Mbps out")
		print_stats(to_return.client_conn_active_stats, "Client CC")
		print_stats(to_return.client_conn_count_stats, "Client CPS")
		print_stats(to_return.server_conn_active_stats, "Server CC")
		print_stats(to_return.server_conn_count_stats, "Server CPS")
	
	return to_return


def write_summary_sheet_header(worksheet):
	ws = worksheet
	ws["A1"] = 'Test Name'
	
	ws["B1"] = 'Sys CPU Avg'
	ws["C1"] = 'Sys CPU Min'
	ws["D1"] = 'Sys CPU Max'
	ws["E1"] = 'Sys CPU StDev'

	ws["F1"] = 'TMM CPU Avg'
	ws["G1"] = 'TMM CPU Min'
	ws["H1"] = 'TMM CPU Max'
	ws["I1"] = 'TMM CPU StDev'

	ws["J1"] = 'Memory Avg'
	ws["K1"] = 'Memory Min'
	ws["L1"] = 'Memory Max'
	ws["M1"] = 'Memory StDev'

	ws["N1"] = 'Client Mb In Avg'
	ws["O1"] = 'Client Mb In Min'
	ws["P1"] = 'Client Mb In Max'
	ws["Q1"] = 'Client Mb In StDev'

	ws["R1"] = 'Client Mb Out Avg'
	ws["S1"] = 'Client Mb Out Min'
	ws["T1"] = 'Client Mb Out Max'
	ws["U1"] = 'Client Mb Out StDev'

	ws["V1"] = 'Server Mb In Avg'
	ws["W1"] = 'Server Mb In Min'
	ws["X1"] = 'Server Mb In Max'
	ws["Y1"] = 'Server Mb In StDev'

	ws["Z1"] = 'Server Mb Out Avg'
	ws["AA1"] = 'Server Mb Out Min'
	ws["AB1"] = 'Server Mb Out Max'
	ws["AC1"] = 'Server Mb Out StDev'

	ws["AD1"] = 'Client CC Avg'
	ws["AE1"] = 'Client CC Min'
	ws["AF1"] = 'Client CC Max'
	ws["AG1"] = 'Client CC StDev'

	ws["AH1"] = 'Client CPS Avg'
	ws["AI1"] = 'Client CPS Min'
	ws["AJ1"] = 'Client CPS Max'
	ws["AK1"] = 'Client CPS StDev'

	ws["AL1"] = 'Server CC Avg'
	ws["AM1"] = 'Server CC Min'
	ws["AN1"] = 'Server CC Max'
	ws["AO1"] = 'Server CC StDev'

	ws["AP1"] = 'Server CPS Avg'
	ws["AQ1"] = 'Server CPS Min'
	ws["AR1"] = 'Server CPS Max'
	ws["AS1"] = 'Server CPS StDev'


def write_summary_sheet_data(perf_data, worksheet, data_name, row_offset):
	"""
	Takes a PerfData, Worksheet, meaningful (to caller) string for identification, integer row offsets.
	Writes the contents of the PerfData to the Worksheet.
	"""
	ws = worksheet
	row = row_offset

	# rows and columns in Excel land are 1-indexed. Account for that in all cell operations.
	ws["A" + str(row)] = data_name
	
	ws["B" + str(row)] = perf_data.sys_cpu_stats.val_mean
	ws["C" + str(row)] = perf_data.sys_cpu_stats.val_min
	ws["D" + str(row)] = perf_data.sys_cpu_stats.val_max
	ws["E" + str(row)] = perf_data.sys_cpu_stats.val_stdev

	ws["F" + str(row)] = perf_data.tmm_cpu_stats.val_mean
	ws["G" + str(row)] = perf_data.tmm_cpu_stats.val_min
	ws["H" + str(row)] = perf_data.tmm_cpu_stats.val_max
	ws["I" + str(row)] = perf_data.tmm_cpu_stats.val_stdev
	
	ws["J" + str(row)] = perf_data.memory_stats.val_mean
	ws["K" + str(row)] = perf_data.memory_stats.val_min
	ws["L" + str(row)] = perf_data.memory_stats.val_max
	ws["M" + str(row)] = perf_data.memory_stats.val_stdev

	ws["N" + str(row)] = perf_data.client_mb_in_stats.val_mean
	ws["O" + str(row)] = perf_data.client_mb_in_stats.val_min
	ws["P" + str(row)] = perf_data.client_mb_in_stats.val_max
	ws["Q" + str(row)] = perf_data.client_mb_in_stats.val_stdev
	
	ws["R" + str(row)] = perf_data.client_mb_out_stats.val_mean
	ws["S" + str(row)] = perf_data.client_mb_out_stats.val_min
	ws["T" + str(row)] = perf_data.client_mb_out_stats.val_max
	ws["U" + str(row)] = perf_data.client_mb_out_stats.val_stdev
	
	ws["V" + str(row)] = perf_data.server_mb_in_stats.val_mean
	ws["W" + str(row)] = perf_data.server_mb_in_stats.val_min
	ws["X" + str(row)] = perf_data.server_mb_in_stats.val_max
	ws["Y" + str(row)] = perf_data.server_mb_in_stats.val_stdev
	
	ws["Z" + str(row)] = perf_data.server_mb_out_stats.val_mean
	ws["AA" + str(row)] = perf_data.server_mb_out_stats.val_min
	ws["AB" + str(row)] = perf_data.server_mb_out_stats.val_max
	ws["AC" + str(row)] = perf_data.server_mb_out_stats.val_stdev
	
	ws["AD" + str(row)] = perf_data.client_conn_active_stats.val_mean
	ws["AE" + str(row)] = perf_data.client_conn_active_stats.val_min
	ws["AF" + str(row)] = perf_data.client_conn_active_stats.val_max
	ws["AG" + str(row)] = perf_data.client_conn_active_stats.val_stdev
	
	ws["AH" + str(row)] = perf_data.client_conn_count_stats.val_mean
	ws["AI" + str(row)] = perf_data.client_conn_count_stats.val_min
	ws["AJ" + str(row)] = perf_data.client_conn_count_stats.val_max
	ws["AK" + str(row)] = perf_data.client_conn_count_stats.val_stdev
	
	ws["AL" + str(row)] = perf_data.server_conn_active_stats.val_mean
	ws["AM" + str(row)] = perf_data.server_conn_active_stats.val_min
	ws["AN" + str(row)] = perf_data.server_conn_active_stats.val_max
	ws["AO" + str(row)] = perf_data.server_conn_active_stats.val_stdev
	
	ws["AP" + str(row)] = perf_data.server_conn_count_stats.val_mean
	ws["AQ" + str(row)] = perf_data.server_conn_count_stats.val_min
	ws["AR" + str(row)] = perf_data.server_conn_count_stats.val_max
	ws["AS" + str(row)] = perf_data.server_conn_count_stats.val_stdev
	

def main():
	
	# compile a list of all workbooks in this tree
	workbook_list = select_workbooks()
	workbook_list.sort()
	if (DEBUG):
		print workbook_list
	
	# create the output workbook
	output_wb = Workbook()
	output_ws = output_wb.active
	output_ws.title = "summary_data"

	write_summary_sheet_header(output_ws)
	foo_count = 2 # rows are 0-indexed in Excel. row 1 is taken by the header.
	for item in workbook_list:
		if (DEBUG):
			print "Processing " + item
		write_summary_sheet_data(process_data(item), output_ws, item, foo_count)
		foo_count += 1
	
	output_wb.save(filename = OUTPUT_WORKBOOK)


if __name__ == "__main__":
	main()
